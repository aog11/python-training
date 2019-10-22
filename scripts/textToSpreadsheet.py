# Chapter 12
# Text Files to Spreadsheet Project

#! python 3

# Importing the needed modules
import os, openpyxl
from openpyxl.styles import Font

# Asking for the location and names of the files
print('Please enter the location of the text files: ',end='')
filesPath = input()
filesPath = os.path.realpath(filesPath)
os.chdir(filesPath)

# Validating that I receive a valid number
validNum = 0
while validNum == 0:
    try:
        print('How many files would you like to work with?: ',end='')
        fileQuantity = int(input())
        validNum = 1
    except ValueError:
        print('Please enter a valid number.')

files = []
for i in range(fileQuantity):
    print('Please enter the name of file #%s: ' %(i+1),end='')
    files.append(input())

# Writing the files to a spreadsheet
wb = openpyxl.Workbook()
sheet = wb.active
for file in range(len(files)):
    try:
        # In case the .txt extension is not specified
        if files[file].endswith('.txt') == False:
            files[file] = files[file] + '.txt'
        textFile = open(files[file],'r')
        # The top row will indicate to which file the content belongs to
        sheet.cell(row=1,column=file+1).font = Font(bold=True)    
        sheet.cell(row=1,column=file+1).value = 'Content of file ' + files[file]
        content = textFile.readlines()
        for c in range(len(content)):
            sheet.cell(row=c+2,column=file+1).value = content[c]
        textFile.close()
    except FileNotFoundError:
        print('File %s does not exist.' %(files[file]))

# Saving the document
wb.save('filestosheet.xlsx')
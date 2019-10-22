# Chapter 12
# Spreadsheet to Text Files project

#! python3

# Importing the needed modules
import os, openpyxl, sys
from openpyxl.utils import get_column_letter

# Asking for the location of the spreadsheet
print('Please enter the location of the spreadsheet: ',end='')
filesPath = input()
filesPath = os.path.realpath(filesPath)
os.chdir(filesPath)

# Asking for the name of the spreadsheet
print('Please enter the name of the spreadsheet: ',end='')
sheetFile = input()

# Validating that the file exists
try:
    if sheetFile.lower().endswith('.xlsx') == False:
        if sheetFile.lower().endswith('.xls') == False:
            if os.path.isfile(sheetFile + '.xlsx'):
                sheetFile = sheetFile + '.xlsx'
            elif os.path.isfile(sheetFile + '.xls'):
                sheetFile = sheetFile + '.xls'
            else:
                raise FileNotFoundError
except FileNotFoundError:
    print('File %s does not exist.' %(sheetFile))
    sys.exit()

wb = openpyxl.load_workbook(sheetFile)
sheet = wb.active

# Writing the contents of the columns to text files
for column in range(1,sheet.max_column+1):
    # Naming the file =  column [column of file] + [file].txt
    textFile = open('column %s of %s.txt' %(get_column_letter(column), os.path.splitext(sheetFile)[0]),'w')
    for row in range(1,sheet.max_row+1):
        textFile.writelines(str(sheet.cell(row=row,column=column).value) + '\n')
    textFile.close()

print('Done.')
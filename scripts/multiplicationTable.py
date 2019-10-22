# Chapter 12
# Multiplication Table Maker Project

#! python3

# Importing the needed modules
import sys, os, openpyxl
from openpyxl.styles import Font

# Verifying that I get the N number
while len(sys.argv) < 2:
    print('Please enter the number to create the table: ', end='')
    try:
        number = int(input())
    except ValueError:
        print('Please enter a valid number.')
        continue
    sys.argv.append(number)
number = int(sys.argv[1])

# Specifying the directory of the file
filePath = ''
os.chdir(filePath)

# Creating the Excel file
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Table of %s' %(str(number))

# Formating the row lables in bold
for rowH in range(1,number + 1):
    sheet.cell(row=rowH+1,column=1).font = Font(bold=True)
    sheet.cell(row=rowH+1,column=1).value = rowH
    # Formatting the column lables in bold
    for colH in range(1,number + 1):
        sheet.cell(row=1,column=colH+1).font = Font(bold=True)
        sheet.cell(row=1,column=colH+1).value = colH
        # Doing the multiplication and assigning it to a cell
        sheet.cell(row=rowH+1, column=colH+1).value = rowH*colH

# Saving the file
wb.save('tableOf%s.xlsx' %(str(number)))